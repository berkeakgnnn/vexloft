#!/usr/bin/env python3
"""
Vexloft Data — demo dashboard generator.

Single source of truth for the four portfolio demo dashboards.
Produces, from the same in-memory sample data:
  - lib/demo-data/data/*.json   -> consumed by the Next.js web dashboards
  - public/demo/excel/*.xlsx     -> real, styled Excel workbooks (download + OneDrive embed)

All data is fictional (invented companies) and clearly presented as sample work.
Run:  python3 scripts/demo-data/generate.py
"""

from __future__ import annotations

import json
import os
import random
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --------------------------------------------------------------------------------------
# Paths
# --------------------------------------------------------------------------------------
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
JSON_DIR = os.path.join(ROOT, "lib", "demo-data", "data")
XLSX_DIR = os.path.join(ROOT, "public", "demo", "excel")
os.makedirs(JSON_DIR, exist_ok=True)
os.makedirs(XLSX_DIR, exist_ok=True)

# --------------------------------------------------------------------------------------
# Brand palette (Vexloft) — reused across web + excel
# --------------------------------------------------------------------------------------
INDIGO = "4338CA"
VIOLET = "7C3AED"
CYAN = "06B6D4"
AMBER = "F59E0B"
GREEN = "10B981"
RED = "EF4444"
BLUE = "3B82F6"
DARK = "0A0F1E"
SLATE = "334155"
LIGHT = "F1F5F9"
CARD_BORDER = "E2E8F0"
WHITE = "FFFFFF"

CHART_PALETTE = [INDIGO, VIOLET, CYAN, AMBER, GREEN, BLUE, RED, "8B5CF6"]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# --------------------------------------------------------------------------------------
# Excel styling helpers
# --------------------------------------------------------------------------------------
THIN = Side(style="thin", color=CARD_BORDER)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def title_band(ws, company: str, title: str, last_col: int = 16) -> None:
    """Dark indigo hero band across the top of a dashboard sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=last_col)
    c = ws.cell(row=1, column=1)
    c.value = f"  {company.upper()}"
    c.font = Font(name="Calibri", size=22, bold=True, color=WHITE)
    c.fill = fill(INDIGO)
    c.alignment = Alignment(vertical="center", horizontal="left")
    for col in range(1, last_col + 1):
        ws.cell(row=1, column=col).fill = fill(INDIGO)
        ws.cell(row=2, column=col).fill = fill(INDIGO)
        ws.cell(row=3, column=col).fill = fill(INDIGO)
    sub = ws.cell(row=4, column=1)
    sub.value = f"  {title}   ·   Sample dashboard by Atakan Harman — Vexloft Data"
    sub.font = Font(name="Calibri", size=10, italic=True, color=SLATE)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 20


def kpi_card(ws, row: int, col: int, label: str, value: str, accent: str,
             width: int = 3) -> None:
    """A 3-col x 3-row KPI card: accent bar on top, big number, small label."""
    # accent bar
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
    for c in range(col, col + width):
        ws.cell(row=row, column=c).fill = fill(accent)
    ws.row_dimensions[row].height = 5
    # value
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + width - 1)
    v = ws.cell(row=row + 1, column=col)
    v.value = value
    v.font = Font(name="Calibri", size=18, bold=True, color=DARK)
    v.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    v.fill = fill(WHITE)
    # label
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + width - 1)
    lb = ws.cell(row=row + 2, column=col)
    lb.value = label
    lb.font = Font(name="Calibri", size=9, color=SLATE)
    lb.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lb.fill = fill(WHITE)
    # borders
    for r in range(row, row + 3):
        for c in range(col, col + width):
            cell = ws.cell(row=r, column=c)
            if r != row:
                cell.fill = fill(WHITE)
            cell.border = BORDER_ALL
    ws.row_dimensions[row + 1].height = 26
    ws.row_dimensions[row + 2].height = 16


def color_series(chart, indices=None) -> None:
    """Apply brand palette to chart series."""
    for i, s in enumerate(chart.series):
        color = CHART_PALETTE[i % len(CHART_PALETTE)]
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color


def color_pie_points(chart, n: int) -> None:
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    pts = []
    for i in range(n):
        gp = GraphicalProperties(solidFill=CHART_PALETTE[i % len(CHART_PALETTE)])
        dp = DataPoint(idx=i, spPr=gp)
        pts.append(dp)
    if chart.series:
        chart.series[0].data_points = pts


def style_table_sheet(ws, header_row: int, first_row: int, last_row: int,
                      last_col: int, name: str) -> None:
    """Style a data table with header fill + Excel Table (autofilter)."""
    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = fill(SLATE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    for r in range(first_row, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if r % 2 == 0:
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill("F8FAFC")
    ref = f"{get_column_letter(1)}{header_row}:{get_column_letter(last_col)}{last_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)


def set_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# --------------------------------------------------------------------------------------
# Formatting helpers for JSON KPI values
# --------------------------------------------------------------------------------------
def eur(n: float) -> str:
    return "€" + f"{round(n):,}"


def eurk(n: float) -> str:
    if abs(n) >= 1_000_000:
        return "€" + f"{n/1_000_000:.2f}M"
    if abs(n) >= 1_000:
        return "€" + f"{n/1_000:.0f}K"
    return "€" + f"{round(n):,}"


def pct(n: float) -> str:
    return f"{n:.1f}%"


# ======================================================================================
# 1. SALES — NovaRetail
# ======================================================================================
def build_sales() -> dict[str, Any]:
    rng = random.Random(101)
    categories = ["Electronics", "Apparel", "Home & Living", "Beauty", "Sports"]
    regions = ["North", "South", "East", "West"]
    cat_weight = {"Electronics": 0.34, "Apparel": 0.24, "Home & Living": 0.18,
                  "Beauty": 0.14, "Sports": 0.10}
    reg_weight = {"North": 0.31, "South": 0.27, "East": 0.24, "West": 0.18}
    # seasonality multipliers (Q4 peak)
    season = [0.82, 0.80, 0.92, 0.95, 1.0, 1.03, 0.98, 0.96, 1.05, 1.12, 1.28, 1.35]
    base_month = 240_000
    margin = {"Electronics": 0.22, "Apparel": 0.42, "Home & Living": 0.35,
              "Beauty": 0.48, "Sports": 0.38}

    rows = []  # granular: month x category x region
    for mi, m in enumerate(MONTHS):
        month_total = base_month * season[mi]
        for cat in categories:
            for reg in regions:
                rev = month_total * cat_weight[cat] * reg_weight[reg]
                rev *= rng.uniform(0.88, 1.12)
                rev = round(rev)
                cost = round(rev * (1 - margin[cat]) * rng.uniform(0.97, 1.03))
                units = round(rev / rng.uniform(38, 120))
                rows.append({
                    "month": m, "monthIndex": mi, "category": cat, "region": reg,
                    "revenue": rev, "cost": cost, "profit": rev - cost, "units": units,
                })

    total_rev = sum(r["revenue"] for r in rows)
    total_profit = sum(r["profit"] for r in rows)
    total_units = sum(r["units"] for r in rows)
    target_total = round(total_rev * 1.06)  # slightly missed target -> realistic

    rev_by_month = [sum(r["revenue"] for r in rows if r["monthIndex"] == i) for i in range(12)]
    profit_by_month = [sum(r["profit"] for r in rows if r["monthIndex"] == i) for i in range(12)]
    target_by_month = [round(v * rng.uniform(1.0, 1.12)) for v in rev_by_month]
    rev_by_cat = [{"name": c, "value": sum(r["revenue"] for r in rows if r["category"] == c)}
                  for c in categories]
    rev_by_reg = [{"name": g, "value": sum(r["revenue"] for r in rows if r["region"] == g)}
                  for g in regions]

    products = [
        ("Aurora Wireless Earbuds", "Electronics"), ("Nimbus 4K Monitor", "Electronics"),
        ("Vertex Running Shoes", "Sports"), ("Lumen Skincare Set", "Beauty"),
        ("Cloudline Duvet", "Home & Living"), ("Pulse Smartwatch", "Electronics"),
        ("Terra Yoga Mat", "Sports"), ("Silk Serum Pro", "Beauty"),
        ("Metro Denim Jacket", "Apparel"), ("Halo Table Lamp", "Home & Living"),
    ]
    top_products = []
    for name, cat in products:
        val = round(total_rev * cat_weight[cat] * rng.uniform(0.02, 0.06))
        top_products.append({"name": name, "category": cat, "value": val})
    top_products.sort(key=lambda x: -x["value"])

    kpis = [
        {"label": "Total Revenue", "value": eurk(total_rev), "accent": INDIGO,
         "delta": "+12.4%", "deltaUp": True, "sub": "vs prior year"},
        {"label": "Gross Profit", "value": eurk(total_profit), "accent": VIOLET,
         "delta": "+9.1%", "deltaUp": True, "sub": "vs prior year"},
        {"label": "Profit Margin", "value": pct(total_profit / total_rev * 100), "accent": CYAN,
         "delta": "+1.8 pts", "deltaUp": True, "sub": "blended"},
        {"label": "Target Achievement", "value": pct(total_rev / target_total * 100),
         "accent": AMBER, "delta": "-5.7%", "deltaUp": False, "sub": "vs annual target"},
        {"label": "Units Sold", "value": f"{total_units:,}", "accent": GREEN,
         "delta": "+8.0%", "deltaUp": True, "sub": "full year"},
        {"label": "Avg Order Value", "value": eur(total_rev / total_units * rng.uniform(2.6, 3.2)),
         "accent": BLUE, "delta": "+3.2%", "deltaUp": True, "sub": "per order"},
    ]

    data = {
        "id": "sales", "slug": "sales", "company": "NovaRetail",
        "title": "Sales Analytics & KPI Dashboard", "currency": "€",
        "period": "FY 2025", "months": MONTHS,
        "categories": categories, "regions": regions,
        "kpis": kpis,
        "series": {
            "revenueByMonth": rev_by_month, "profitByMonth": profit_by_month,
            "targetByMonth": target_by_month,
            "revenueByCategory": rev_by_cat, "revenueByRegion": rev_by_reg,
            "topProducts": top_products[:8],
        },
        "rows": rows,
    }
    _excel_sales(data)
    return data


def _excel_sales(d: dict[str, Any]) -> None:
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    set_widths(dash, {i: 11 for i in range(1, 17)})
    title_band(dash, d["company"], d["title"])

    cards = d["kpis"]
    positions = [(6, 1), (6, 4), (6, 7), (6, 10), (6, 13), (10, 1)]
    # simpler: 5 across the top
    positions = [(6, 1 + i * 3) for i in range(5)]
    for (r, c), k in zip(positions, cards[:5]):
        kpi_card(dash, r, c, k["label"], k["value"], k["accent"])

    # model sheet
    md = wb.create_sheet("Model")
    md.sheet_state = "hidden"
    md["A1"] = "Month"
    md["B1"] = "Revenue"
    md["C1"] = "Target"
    md["D1"] = "Profit"
    for i, m in enumerate(d["months"]):
        md.cell(row=2 + i, column=1, value=m)
        md.cell(row=2 + i, column=2, value=d["series"]["revenueByMonth"][i])
        md.cell(row=2 + i, column=3, value=d["series"]["targetByMonth"][i])
        md.cell(row=2 + i, column=4, value=d["series"]["profitByMonth"][i])
    md["F1"] = "Category"
    md["G1"] = "Revenue"
    for i, cat in enumerate(d["series"]["revenueByCategory"]):
        md.cell(row=2 + i, column=6, value=cat["name"])
        md.cell(row=2 + i, column=7, value=cat["value"])
    md["I1"] = "Region"
    md["J1"] = "Revenue"
    for i, reg in enumerate(d["series"]["revenueByRegion"]):
        md.cell(row=2 + i, column=9, value=reg["name"])
        md.cell(row=2 + i, column=10, value=reg["value"])

    # Chart 1: Revenue vs Target by month (combo-ish: bars + line)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Revenue vs Target by Month"
    bar.height = 7.5
    bar.width = 16
    data_ref = Reference(md, min_col=2, min_row=1, max_row=13)
    cats_ref = Reference(md, min_col=1, min_row=2, max_row=13)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    line = LineChart()
    line_ref = Reference(md, min_col=3, min_row=1, max_row=13)
    line.add_data(line_ref, titles_from_data=True)
    bar += line
    color_series(bar)
    bar.series[0].graphicalProperties.solidFill = INDIGO
    dash.add_chart(bar, "A10")

    # Chart 2: Revenue by Category (doughnut)
    dough = DoughnutChart()
    dough.title = "Revenue by Category"
    dough.height = 7.5
    dough.width = 8
    d_ref = Reference(md, min_col=7, min_row=1, max_row=1 + len(d["categories"]))
    d_cats = Reference(md, min_col=6, min_row=2, max_row=1 + len(d["categories"]))
    dough.add_data(d_ref, titles_from_data=True)
    dough.set_categories(d_cats)
    color_pie_points(dough, len(d["categories"]))
    dash.add_chart(dough, "J10")

    # Chart 3: Revenue by Region (bar)
    rbar = BarChart()
    rbar.type = "bar"
    rbar.title = "Revenue by Region"
    rbar.height = 7.5
    rbar.width = 8
    r_ref = Reference(md, min_col=10, min_row=1, max_row=1 + len(d["regions"]))
    r_cats = Reference(md, min_col=9, min_row=2, max_row=1 + len(d["regions"]))
    rbar.add_data(r_ref, titles_from_data=True)
    rbar.set_categories(r_cats)
    rbar.series[0].graphicalProperties.solidFill = VIOLET
    rbar.legend = None
    dash.add_chart(rbar, "A26")

    # Data sheet (granular)
    ds = wb.create_sheet("Sales Data")
    ds.sheet_view.showGridLines = False
    headers = ["Month", "Category", "Region", "Units", "Revenue (€)", "Cost (€)", "Profit (€)"]
    for c, h in enumerate(headers, start=1):
        ds.cell(row=1, column=c, value=h)
    for i, row in enumerate(d["rows"], start=2):
        ds.cell(row=i, column=1, value=row["month"])
        ds.cell(row=i, column=2, value=row["category"])
        ds.cell(row=i, column=3, value=row["region"])
        ds.cell(row=i, column=4, value=row["units"])
        ds.cell(row=i, column=5, value=row["revenue"])
        ds.cell(row=i, column=6, value=row["cost"])
        ds.cell(row=i, column=7, value=row["profit"])
    last = len(d["rows"]) + 1
    for col in (5, 6, 7):
        for r in range(2, last + 1):
            ds.cell(row=r, column=col).number_format = "#,##0"
    style_table_sheet(ds, 1, 2, last, len(headers), "SalesData")
    set_widths(ds, {1: 10, 2: 16, 3: 10, 4: 10, 5: 14, 6: 14, 7: 14})
    ds.freeze_panes = "A2"
    ds.column_dimensions["G"].width = 14
    # profit color scale
    ds.conditional_formatting.add(
        f"G2:G{last}",
        ColorScaleRule(start_type="min", start_color=RED,
                       mid_type="percentile", mid_value=50, mid_color=AMBER,
                       end_type="max", end_color=GREEN))

    _finish(wb, "sales")


# ======================================================================================
# 2. INVENTORY — AeroParts Ltd
# ======================================================================================
def build_inventory() -> dict[str, Any]:
    rng = random.Random(202)
    cats = ["Fasteners", "Avionics", "Hydraulics", "Engine Parts", "Interior"]
    warehouses = ["Hub-A", "Hub-B", "Hub-C"]
    part_names = {
        "Fasteners": ["Titanium Bolt", "Lock Nut", "Rivet Set", "Hex Screw", "Washer Kit"],
        "Avionics": ["Nav Transceiver", "Altimeter Module", "GPS Unit", "Comm Radio", "Sensor Board"],
        "Hydraulics": ["Actuator Cyl.", "Pump Assembly", "Valve Block", "Hose Line", "Seal Kit"],
        "Engine Parts": ["Turbine Blade", "Fuel Injector", "Bearing Ring", "Compressor Disk", "Igniter Plug"],
        "Interior": ["Seat Rail", "Panel Trim", "Cabin Light", "Tray Table", "Overhead Bin Latch"],
    }
    rows = []
    sku_no = 1000
    for cat in cats:
        for name in part_names[cat]:
            for _ in range(2):
                sku_no += 7
                onhand = rng.randint(0, 480)
                reorder = rng.randint(60, 180)
                unit_cost = round(rng.uniform(12, 2400), 2)
                usage = rng.randint(15, 140)
                lead = rng.randint(5, 45)
                if onhand == 0:
                    status = "Out of Stock"
                elif onhand < reorder:
                    status = "Reorder"
                elif onhand < reorder * 1.4:
                    status = "Low"
                else:
                    status = "In Stock"
                rows.append({
                    "sku": f"AP-{sku_no}", "part": name, "category": cat,
                    "warehouse": rng.choice(warehouses), "onHand": onhand,
                    "reorderPoint": reorder, "unitCost": unit_cost,
                    "value": round(onhand * unit_cost), "monthlyUsage": usage,
                    "leadTimeDays": lead, "status": status,
                })
    total_skus = len(rows)
    total_value = sum(r["value"] for r in rows)
    below = sum(1 for r in rows if r["status"] in ("Reorder", "Out of Stock"))
    oos = sum(1 for r in rows if r["status"] == "Out of Stock")
    avg_lead = sum(r["leadTimeDays"] for r in rows) / total_skus
    annual_usage_val = sum(r["monthlyUsage"] * r["unitCost"] * 12 for r in rows)
    turnover = annual_usage_val / total_value

    value_by_cat = [{"name": c, "value": sum(r["value"] for r in rows if r["category"] == c)}
                    for c in cats]
    status_order = ["In Stock", "Low", "Reorder", "Out of Stock"]
    status_dist = [{"name": s, "value": sum(1 for r in rows if r["status"] == s)}
                   for s in status_order]
    top_value = sorted(rows, key=lambda r: -r["value"])[:8]
    top_value_series = [{"name": r["part"] + " " + r["sku"][-3:], "value": r["value"]}
                        for r in top_value]

    kpis = [
        {"label": "Total SKUs", "value": f"{total_skus}", "accent": INDIGO,
         "delta": "50 active", "deltaUp": True, "sub": "tracked parts"},
        {"label": "Inventory Value", "value": eurk(total_value), "accent": VIOLET,
         "delta": "+4.6%", "deltaUp": True, "sub": "on-hand at cost"},
        {"label": "Below Reorder", "value": f"{below}", "accent": AMBER,
         "delta": f"{oos} out of stock", "deltaUp": False, "sub": "needs action"},
        {"label": "Inventory Turnover", "value": f"{turnover:.1f}x", "accent": CYAN,
         "delta": "+0.3x", "deltaUp": True, "sub": "annualised"},
        {"label": "Avg Lead Time", "value": f"{avg_lead:.0f} days", "accent": GREEN,
         "delta": "-2 days", "deltaUp": True, "sub": "supplier avg"},
    ]
    data = {
        "id": "inventory", "slug": "inventory", "company": "AeroParts Ltd",
        "title": "Inventory Management Dashboard", "currency": "€",
        "period": "Live snapshot", "categories": cats, "warehouses": warehouses,
        "statusOrder": status_order,
        "kpis": kpis,
        "series": {
            "valueByCategory": value_by_cat, "statusDistribution": status_dist,
            "topValueSkus": top_value_series,
        },
        "rows": rows,
    }
    _excel_inventory(data)
    return data


def _excel_inventory(d: dict[str, Any]) -> None:
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    set_widths(dash, {i: 11 for i in range(1, 17)})
    title_band(dash, d["company"], d["title"])
    for i, k in enumerate(d["kpis"][:5]):
        kpi_card(dash, 6, 1 + i * 3, k["label"], k["value"], k["accent"])

    md = wb.create_sheet("Model")
    md.sheet_state = "hidden"
    md["A1"] = "Category"
    md["B1"] = "Value"
    for i, x in enumerate(d["series"]["valueByCategory"]):
        md.cell(row=2 + i, column=1, value=x["name"])
        md.cell(row=2 + i, column=2, value=x["value"])
    md["D1"] = "Status"
    md["E1"] = "Count"
    for i, x in enumerate(d["series"]["statusDistribution"]):
        md.cell(row=2 + i, column=4, value=x["name"])
        md.cell(row=2 + i, column=5, value=x["value"])
    md["G1"] = "SKU"
    md["H1"] = "Value"
    for i, x in enumerate(d["series"]["topValueSkus"]):
        md.cell(row=2 + i, column=7, value=x["name"])
        md.cell(row=2 + i, column=8, value=x["value"])

    cbar = BarChart()
    cbar.type = "col"
    cbar.title = "Inventory Value by Category"
    cbar.height = 7.5
    cbar.width = 9
    cbar.add_data(Reference(md, min_col=2, min_row=1, max_row=1 + len(d["categories"])),
                  titles_from_data=True)
    cbar.set_categories(Reference(md, min_col=1, min_row=2, max_row=1 + len(d["categories"])))
    cbar.series[0].graphicalProperties.solidFill = INDIGO
    cbar.legend = None
    dash.add_chart(cbar, "A10")

    dough = DoughnutChart()
    dough.title = "Stock Status Distribution"
    dough.height = 7.5
    dough.width = 7
    dough.add_data(Reference(md, min_col=5, min_row=1, max_row=1 + len(d["statusOrder"])),
                   titles_from_data=True)
    dough.set_categories(Reference(md, min_col=4, min_row=2, max_row=1 + len(d["statusOrder"])))
    color_pie_points(dough, len(d["statusOrder"]))
    dash.add_chart(dough, "H10")

    tbar = BarChart()
    tbar.type = "bar"
    tbar.title = "Top SKUs by Inventory Value"
    tbar.height = 7.5
    tbar.width = 9
    tbar.add_data(Reference(md, min_col=8, min_row=1, max_row=1 + 8), titles_from_data=True)
    tbar.set_categories(Reference(md, min_col=7, min_row=2, max_row=1 + 8))
    tbar.series[0].graphicalProperties.solidFill = CYAN
    tbar.legend = None
    dash.add_chart(tbar, "N10")

    ds = wb.create_sheet("Inventory Data")
    ds.sheet_view.showGridLines = False
    headers = ["SKU", "Part", "Category", "Warehouse", "On Hand", "Reorder Pt",
               "Unit Cost (€)", "Value (€)", "Monthly Usage", "Lead Time (d)", "Status"]
    for c, h in enumerate(headers, start=1):
        ds.cell(row=1, column=c, value=h)
    for i, r in enumerate(d["rows"], start=2):
        vals = [r["sku"], r["part"], r["category"], r["warehouse"], r["onHand"],
                r["reorderPoint"], r["unitCost"], r["value"], r["monthlyUsage"],
                r["leadTimeDays"], r["status"]]
        for c, v in enumerate(vals, start=1):
            ds.cell(row=i, column=c, value=v)
    last = len(d["rows"]) + 1
    for r in range(2, last + 1):
        ds.cell(row=r, column=7).number_format = "#,##0.00"
        ds.cell(row=r, column=8).number_format = "#,##0"
    style_table_sheet(ds, 1, 2, last, len(headers), "InventoryData")
    set_widths(ds, {1: 10, 2: 18, 3: 14, 4: 11, 5: 10, 6: 11, 7: 12, 8: 12, 9: 13, 10: 12, 11: 13})
    ds.freeze_panes = "A2"
    # status conditional formatting
    ds.conditional_formatting.add(
        f"K2:K{last}", CellIsRule(operator="equal", formula=['"Out of Stock"'],
                                  fill=fill("FECACA"), font=Font(color="991B1B", bold=True)))
    ds.conditional_formatting.add(
        f"K2:K{last}", CellIsRule(operator="equal", formula=['"Reorder"'],
                                  fill=fill("FED7AA"), font=Font(color="9A3412", bold=True)))
    ds.conditional_formatting.add(
        f"K2:K{last}", CellIsRule(operator="equal", formula=['"Low"'],
                                  fill=fill("FEF08A"), font=Font(color="854D0E")))
    ds.conditional_formatting.add(
        f"K2:K{last}", CellIsRule(operator="equal", formula=['"In Stock"'],
                                  fill=fill("BBF7D0"), font=Font(color="166534")))
    _finish(wb, "inventory")


# ======================================================================================
# 3. PROCUREMENT — Meridian Manufacturing
# ======================================================================================
def build_procurement() -> dict[str, Any]:
    rng = random.Random(303)
    cats = ["Raw Materials", "Components", "Packaging", "MRO", "Logistics"]
    suppliers = ["Kraftwerk GmbH", "Nordic Steel Co", "Apex Components",
                 "PoliPack SA", "Volta Logistics", "Summit MRO", "Delta Alloys", "BluePine Ltd"]
    buyers = ["A. Ozdemir", "M. Chen", "L. Rossi", "K. Novak"]
    statuses = ["Received", "Open", "Overdue"]
    rows = []
    po_no = 4800
    for _ in range(42):
        po_no += 3
        cat = rng.choice(cats)
        sup = rng.choice(suppliers)
        value = round(rng.uniform(3_500, 145_000))
        status = rng.choices(statuses, weights=[0.6, 0.28, 0.12])[0]
        lead = rng.randint(7, 60)
        on_time = "Yes" if (status == "Received" and rng.random() > 0.22) else ("No" if status == "Overdue" else "-")
        savings = round(value * rng.uniform(0, 0.09))
        rows.append({
            "po": f"PO-{po_no}", "supplier": sup, "category": cat,
            "buyer": rng.choice(buyers), "value": value, "status": status,
            "leadTimeDays": lead, "onTime": on_time, "savings": savings,
            "monthIndex": rng.randint(0, 11),
        })
    total_spend = sum(r["value"] for r in rows)
    open_po = sum(1 for r in rows if r["status"] == "Open")
    overdue = sum(1 for r in rows if r["status"] == "Overdue")
    savings_ytd = sum(r["savings"] for r in rows)
    received = [r for r in rows if r["status"] == "Received"]
    otd = (sum(1 for r in received if r["onTime"] == "Yes") / max(len(received), 1)) * 100

    # supplier scorecard
    scorecard = []
    for sup in suppliers:
        srows = [r for r in rows if r["supplier"] == sup]
        spend = sum(r["value"] for r in srows)
        s_otd = rng.uniform(72, 99)
        s_lead = sum(r["leadTimeDays"] for r in srows) / max(len(srows), 1)
        quality = rng.uniform(78, 99)
        price = rng.uniform(80, 98)
        overall = round((s_otd * 0.35 + quality * 0.35 + price * 0.30), 1)
        scorecard.append({
            "supplier": sup, "spend": spend, "otd": round(s_otd, 1),
            "leadTime": round(s_lead), "quality": round(quality, 1),
            "priceIndex": round(price, 1), "overall": overall,
            "rating": "A" if overall >= 90 else ("B" if overall >= 82 else "C"),
        })
    scorecard.sort(key=lambda x: -x["overall"])

    spend_by_cat = [{"name": c, "value": sum(r["value"] for r in rows if r["category"] == c)}
                    for c in cats]
    spend_by_sup = sorted(
        [{"name": s, "value": sum(r["value"] for r in rows if r["supplier"] == s)}
         for s in suppliers], key=lambda x: -x["value"])
    status_dist = [{"name": s, "value": sum(1 for r in rows if r["status"] == s)}
                   for s in statuses]
    spend_by_month = [sum(r["value"] for r in rows if r["monthIndex"] == i) for i in range(12)]

    kpis = [
        {"label": "Total Spend", "value": eurk(total_spend), "accent": INDIGO,
         "delta": "YTD", "deltaUp": True, "sub": "across all POs"},
        {"label": "Open POs", "value": f"{open_po}", "accent": VIOLET,
         "delta": f"{overdue} overdue", "deltaUp": False, "sub": "in progress"},
        {"label": "Avg On-Time Delivery", "value": pct(otd), "accent": GREEN,
         "delta": "+3.1 pts", "deltaUp": True, "sub": "received orders"},
        {"label": "Savings YTD", "value": eurk(savings_ytd), "accent": CYAN,
         "delta": pct(savings_ytd / total_spend * 100), "deltaUp": True, "sub": "vs baseline"},
        {"label": "Active Suppliers", "value": f"{len(suppliers)}", "accent": AMBER,
         "delta": "3 A-rated", "deltaUp": True, "sub": "qualified"},
    ]
    data = {
        "id": "procurement", "slug": "procurement", "company": "Meridian Manufacturing",
        "title": "Procurement & Supplier Dashboard", "currency": "€",
        "period": "FY 2025", "categories": cats, "suppliers": suppliers, "statuses": statuses,
        "kpis": kpis,
        "series": {
            "spendByCategory": spend_by_cat, "spendBySupplier": spend_by_sup,
            "statusDistribution": status_dist, "spendByMonth": spend_by_month,
            "months": MONTHS,
        },
        "scorecard": scorecard, "rows": rows,
    }
    _excel_procurement(data)
    return data


def _excel_procurement(d: dict[str, Any]) -> None:
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    set_widths(dash, {i: 11 for i in range(1, 17)})
    title_band(dash, d["company"], d["title"])
    for i, k in enumerate(d["kpis"][:5]):
        kpi_card(dash, 6, 1 + i * 3, k["label"], k["value"], k["accent"])

    md = wb.create_sheet("Model")
    md.sheet_state = "hidden"
    md["A1"] = "Category"
    md["B1"] = "Spend"
    for i, x in enumerate(d["series"]["spendByCategory"]):
        md.cell(row=2 + i, column=1, value=x["name"])
        md.cell(row=2 + i, column=2, value=x["value"])
    md["D1"] = "Month"
    md["E1"] = "Spend"
    for i, m in enumerate(d["series"]["months"]):
        md.cell(row=2 + i, column=4, value=m)
        md.cell(row=2 + i, column=5, value=d["series"]["spendByMonth"][i])
    md["G1"] = "Supplier"
    md["H1"] = "OTD"
    for i, s in enumerate(d["scorecard"]):
        md.cell(row=2 + i, column=7, value=s["supplier"])
        md.cell(row=2 + i, column=8, value=s["otd"])

    cbar = BarChart()
    cbar.type = "col"
    cbar.title = "Spend by Category"
    cbar.height = 7.5
    cbar.width = 9
    cbar.add_data(Reference(md, min_col=2, min_row=1, max_row=1 + len(d["categories"])),
                  titles_from_data=True)
    cbar.set_categories(Reference(md, min_col=1, min_row=2, max_row=1 + len(d["categories"])))
    cbar.series[0].graphicalProperties.solidFill = INDIGO
    cbar.legend = None
    dash.add_chart(cbar, "A10")

    line = LineChart()
    line.title = "Monthly Spend Trend"
    line.height = 7.5
    line.width = 9
    line.add_data(Reference(md, min_col=5, min_row=1, max_row=13), titles_from_data=True)
    line.set_categories(Reference(md, min_col=4, min_row=2, max_row=13))
    line.series[0].graphicalProperties.line.solidFill = VIOLET
    line.series[0].graphicalProperties.line.width = 28000
    line.series[0].marker = Marker(symbol="circle", size=5)
    line.legend = None
    dash.add_chart(line, "H10")

    obar = BarChart()
    obar.type = "bar"
    obar.title = "Supplier On-Time Delivery %"
    obar.height = 7.5
    obar.width = 9
    obar.add_data(Reference(md, min_col=8, min_row=1, max_row=1 + len(d["suppliers"])),
                  titles_from_data=True)
    obar.set_categories(Reference(md, min_col=7, min_row=2, max_row=1 + len(d["suppliers"])))
    obar.series[0].graphicalProperties.solidFill = GREEN
    obar.legend = None
    dash.add_chart(obar, "N10")

    # Scorecard sheet
    sc = wb.create_sheet("Supplier Scorecard")
    sc.sheet_view.showGridLines = False
    headers = ["Supplier", "Spend (€)", "OTD %", "Avg Lead (d)", "Quality %",
               "Price Index", "Overall", "Rating"]
    for c, h in enumerate(headers, start=1):
        sc.cell(row=1, column=c, value=h)
    for i, s in enumerate(d["scorecard"], start=2):
        vals = [s["supplier"], s["spend"], s["otd"], s["leadTime"], s["quality"],
                s["priceIndex"], s["overall"], s["rating"]]
        for c, v in enumerate(vals, start=1):
            sc.cell(row=i, column=c, value=v)
        sc.cell(row=i, column=2).number_format = "#,##0"
    last = len(d["scorecard"]) + 1
    style_table_sheet(sc, 1, 2, last, len(headers), "SupplierScorecard")
    set_widths(sc, {1: 20, 2: 13, 3: 9, 4: 12, 5: 11, 6: 12, 7: 10, 8: 8})
    sc.conditional_formatting.add(
        f"G2:G{last}", ColorScaleRule(start_type="num", start_value=75, start_color=RED,
                                      mid_type="num", mid_value=85, mid_color=AMBER,
                                      end_type="num", end_value=95, end_color=GREEN))

    # PO data sheet
    ds = wb.create_sheet("Purchase Orders")
    ds.sheet_view.showGridLines = False
    headers = ["PO", "Supplier", "Category", "Buyer", "Value (€)", "Status",
               "Lead Time (d)", "On Time", "Savings (€)"]
    for c, h in enumerate(headers, start=1):
        ds.cell(row=1, column=c, value=h)
    for i, r in enumerate(d["rows"], start=2):
        vals = [r["po"], r["supplier"], r["category"], r["buyer"], r["value"],
                r["status"], r["leadTimeDays"], r["onTime"], r["savings"]]
        for c, v in enumerate(vals, start=1):
            ds.cell(row=i, column=c, value=v)
        ds.cell(row=i, column=5).number_format = "#,##0"
        ds.cell(row=i, column=9).number_format = "#,##0"
    last = len(d["rows"]) + 1
    style_table_sheet(ds, 1, 2, last, len(headers), "PurchaseOrders")
    set_widths(ds, {1: 10, 2: 18, 3: 15, 4: 12, 5: 12, 6: 11, 7: 12, 8: 9, 9: 12})
    ds.freeze_panes = "A2"
    ds.conditional_formatting.add(
        f"F2:F{last}", CellIsRule(operator="equal", formula=['"Overdue"'],
                                  fill=fill("FECACA"), font=Font(color="991B1B", bold=True)))
    ds.conditional_formatting.add(
        f"F2:F{last}", CellIsRule(operator="equal", formula=['"Open"'],
                                  fill=fill("FEF08A"), font=Font(color="854D0E")))
    ds.conditional_formatting.add(
        f"F2:F{last}", CellIsRule(operator="equal", formula=['"Received"'],
                                  fill=fill("BBF7D0"), font=Font(color="166534")))
    _finish(wb, "procurement")


# ======================================================================================
# 4. EXECUTIVE — Vantage Group
# ======================================================================================
def build_executive() -> dict[str, Any]:
    rng = random.Random(404)
    season = [0.86, 0.88, 0.95, 0.98, 1.02, 1.05, 1.0, 0.99, 1.06, 1.1, 1.18, 1.24]
    revenue = [round(1_050_000 * s * rng.uniform(0.97, 1.03)) for s in season]
    ebitda = [round(r * rng.uniform(0.17, 0.23)) for r in revenue]
    departments = ["Retail", "Wholesale", "Digital", "Services"]
    dept_perf = []
    for dep in departments:
        rev = round(sum(revenue) * rng.uniform(0.15, 0.35))
        growth = round(rng.uniform(-4, 22), 1)
        margin = round(rng.uniform(12, 28), 1)
        dept_perf.append({"name": dep, "revenue": rev, "growth": growth, "margin": margin})
    regions = ["Europe", "N. America", "APAC", "MEA"]
    reg_rev = [{"name": r, "value": round(sum(revenue) * w)}
               for r, w in zip(regions, [0.42, 0.28, 0.20, 0.10])]

    total_rev = sum(revenue)
    total_ebitda = sum(ebitda)
    kpis = [
        {"label": "Annual Revenue", "value": eurk(total_rev), "accent": INDIGO,
         "delta": "+14.2%", "deltaUp": True, "sub": "YoY"},
        {"label": "EBITDA", "value": eurk(total_ebitda), "accent": VIOLET,
         "delta": pct(total_ebitda / total_rev * 100), "deltaUp": True, "sub": "margin"},
        {"label": "Win Rate", "value": "34.8%", "accent": CYAN,
         "delta": "+2.4 pts", "deltaUp": True, "sub": "sales pipeline"},
        {"label": "On-Time Delivery", "value": "94.1%", "accent": GREEN,
         "delta": "+1.5 pts", "deltaUp": True, "sub": "operations"},
        {"label": "Headcount", "value": "1,284", "accent": AMBER,
         "delta": "+86 YoY", "deltaUp": True, "sub": "full-time"},
        {"label": "Employee eNPS", "value": "+41", "accent": BLUE,
         "delta": "+6", "deltaUp": True, "sub": "engagement"},
    ]
    data = {
        "id": "executive", "slug": "executive", "company": "Vantage Group",
        "title": "Executive KPI Dashboard", "currency": "€", "period": "FY 2025",
        "months": MONTHS, "departments": departments, "regions": regions,
        "kpis": kpis,
        "series": {
            "revenueByMonth": revenue, "ebitdaByMonth": ebitda,
            "departmentPerformance": dept_perf, "revenueByRegion": reg_rev,
        },
    }
    _excel_executive(data)
    return data


def _excel_executive(d: dict[str, Any]) -> None:
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    set_widths(dash, {i: 11 for i in range(1, 17)})
    title_band(dash, d["company"], d["title"])
    # 6 KPI cards -> two rows of 3? Keep single row of 6 across (18 cols) -> too wide.
    # Use 6 cards across 18 cols but sheet is 16 wide; do 5 across + 1 below.
    for i, k in enumerate(d["kpis"][:5]):
        kpi_card(dash, 6, 1 + i * 3, k["label"], k["value"], k["accent"])
    kpi_card(dash, 6, 16, d["kpis"][5]["label"], d["kpis"][5]["value"], d["kpis"][5]["accent"])

    md = wb.create_sheet("Model")
    md.sheet_state = "hidden"
    md["A1"] = "Month"
    md["B1"] = "Revenue"
    md["C1"] = "EBITDA"
    for i, m in enumerate(d["months"]):
        md.cell(row=2 + i, column=1, value=m)
        md.cell(row=2 + i, column=2, value=d["series"]["revenueByMonth"][i])
        md.cell(row=2 + i, column=3, value=d["series"]["ebitdaByMonth"][i])
    md["E1"] = "Dept"
    md["F1"] = "Revenue"
    for i, x in enumerate(d["series"]["departmentPerformance"]):
        md.cell(row=2 + i, column=5, value=x["name"])
        md.cell(row=2 + i, column=6, value=x["revenue"])
    md["H1"] = "Region"
    md["I1"] = "Revenue"
    for i, x in enumerate(d["series"]["revenueByRegion"]):
        md.cell(row=2 + i, column=8, value=x["name"])
        md.cell(row=2 + i, column=9, value=x["value"])

    combo = BarChart()
    combo.type = "col"
    combo.title = "Revenue vs EBITDA by Month"
    combo.height = 8
    combo.width = 17
    combo.add_data(Reference(md, min_col=2, min_row=1, max_row=13), titles_from_data=True)
    combo.set_categories(Reference(md, min_col=1, min_row=2, max_row=13))
    combo.series[0].graphicalProperties.solidFill = INDIGO
    line = LineChart()
    line.add_data(Reference(md, min_col=3, min_row=1, max_row=13), titles_from_data=True)
    line.series[0].graphicalProperties.line.solidFill = CYAN
    line.series[0].graphicalProperties.line.width = 28000
    combo += line
    dash.add_chart(combo, "A10")

    dbar = BarChart()
    dbar.type = "bar"
    dbar.title = "Revenue by Department"
    dbar.height = 7.5
    dbar.width = 8
    dbar.add_data(Reference(md, min_col=6, min_row=1, max_row=1 + len(d["departments"])),
                  titles_from_data=True)
    dbar.set_categories(Reference(md, min_col=5, min_row=2, max_row=1 + len(d["departments"])))
    dbar.series[0].graphicalProperties.solidFill = VIOLET
    dbar.legend = None
    dash.add_chart(dbar, "A26")

    pie = PieChart()
    pie.title = "Revenue by Region"
    pie.height = 7.5
    pie.width = 8
    pie.add_data(Reference(md, min_col=9, min_row=1, max_row=1 + len(d["regions"])),
                 titles_from_data=True)
    pie.set_categories(Reference(md, min_col=8, min_row=2, max_row=1 + len(d["regions"])))
    color_pie_points(pie, len(d["regions"]))
    dash.add_chart(pie, "J26")

    # Department scorecard
    sc = wb.create_sheet("Department KPIs")
    sc.sheet_view.showGridLines = False
    headers = ["Department", "Revenue (€)", "YoY Growth %", "Margin %"]
    for c, h in enumerate(headers, start=1):
        sc.cell(row=1, column=c, value=h)
    for i, x in enumerate(d["series"]["departmentPerformance"], start=2):
        sc.cell(row=i, column=1, value=x["name"])
        sc.cell(row=i, column=2, value=x["revenue"]).number_format = "#,##0"
        sc.cell(row=i, column=3, value=x["growth"])
        sc.cell(row=i, column=4, value=x["margin"])
    last = len(d["departments"]) + 1
    style_table_sheet(sc, 1, 2, last, len(headers), "DepartmentKPIs")
    set_widths(sc, {1: 16, 2: 14, 3: 14, 4: 12})
    sc.conditional_formatting.add(
        f"C2:C{last}", ColorScaleRule(start_type="min", start_color=RED,
                                      mid_type="num", mid_value=0, mid_color=WHITE,
                                      end_type="max", end_color=GREEN))
    _finish(wb, "executive")


# ======================================================================================
# 5. FINANCIAL — Northwind Foods
# ======================================================================================
def build_financial() -> dict[str, Any]:
    rng = random.Random(505)
    season = [0.9, 0.88, 0.96, 0.98, 1.02, 1.05, 1.0, 0.99, 1.05, 1.08, 1.14, 1.2]
    revenue = [round(880_000 * s * rng.uniform(0.97, 1.03)) for s in season]
    cogs = [round(r * rng.uniform(0.52, 0.58)) for r in revenue]
    opex = [round(r * rng.uniform(0.20, 0.26)) for r in revenue]
    ebitda = [revenue[i] - cogs[i] - opex[i] for i in range(12)]
    cash_in = [round(revenue[i] * rng.uniform(0.9, 1.05)) for i in range(12)]
    cash_out = [round((cogs[i] + opex[i]) * rng.uniform(0.95, 1.08)) for i in range(12)]
    cash_net = [cash_in[i] - cash_out[i] for i in range(12)]
    total_rev = sum(revenue)
    gross = total_rev - sum(cogs)
    total_ebitda = sum(ebitda)
    net_cash = sum(cash_net)
    cash_balance = 1_450_000 + net_cash
    monthly_burn = max(1, round(sum(cash_out) / 12 - sum(cash_in) / 12))
    runway = cash_balance / (abs(monthly_burn) if monthly_burn > 0 else 1)
    expense = [
        {"name": "COGS", "value": sum(cogs)},
        {"name": "Payroll", "value": round(sum(opex) * 0.52)},
        {"name": "Marketing", "value": round(sum(opex) * 0.22)},
        {"name": "Operations", "value": round(sum(opex) * 0.16)},
        {"name": "R&D", "value": round(sum(opex) * 0.10)},
    ]
    depts = ["Production", "Sales", "Logistics", "Admin"]
    budget_actual = []
    for d in depts:
        b = round(total_rev * rng.uniform(0.06, 0.16))
        a = round(b * rng.uniform(0.88, 1.12))
        budget_actual.append({"name": d, "budget": b, "actual": a})

    pos = sum(1 for c in cash_net if c > 0)
    kpis = [
        {"label": "Annual Revenue", "value": eurk(total_rev), "accent": GREEN, "delta": "+11.3%", "deltaUp": True, "sub": "YoY"},
        {"label": "Gross Margin", "value": pct(gross / total_rev * 100), "accent": INDIGO, "delta": "+1.4 pts", "deltaUp": True, "sub": "blended"},
        {"label": "EBITDA", "value": eurk(total_ebitda), "accent": VIOLET, "delta": pct(total_ebitda / total_rev * 100), "deltaUp": True, "sub": "margin"},
        {"label": "Net Cash Flow", "value": eurk(net_cash), "accent": CYAN, "delta": f"{pos}/12 mo +", "deltaUp": True, "sub": "full year"},
        {"label": "Cash Balance", "value": eurk(cash_balance), "accent": AMBER, "delta": f"{eurk(net_cash)} net", "deltaUp": True, "sub": "end of period"},
    ]
    data = {
        "id": "financial", "slug": "financial", "company": "Northwind Foods",
        "title": "Financial & Cash Flow Dashboard", "currency": "€", "period": "FY 2025",
        "months": MONTHS, "kpis": kpis,
        "series": {
            "revenueByMonth": revenue, "ebitdaByMonth": ebitda,
            "cashInByMonth": cash_in, "cashOutByMonth": cash_out, "cashNetByMonth": cash_net,
            "expenseBreakdown": expense, "budgetVsActual": budget_actual,
        },
    }
    _excel_generic(data, "financial", GREEN, [
        ("combo", "Revenue vs EBITDA by Month", "revenueByMonth", "ebitdaByMonth", MONTHS),
        ("doughnut", "Expense Breakdown", "expenseBreakdown", None, None),
        ("bar", "Net Cash Flow by Month", "cashNetByMonth", None, MONTHS),
    ], ("Budget vs Actual", ["Department", "Budget (€)", "Actual (€)", "Variance (€)"],
        [[b["name"], b["budget"], b["actual"], b["actual"] - b["budget"]] for b in budget_actual], [1, 2, 3], 3))
    return data


# ======================================================================================
# 6. MARKETING — Pulse Media
# ======================================================================================
def build_marketing() -> dict[str, Any]:
    rng = random.Random(606)
    chans = ["Google Ads", "Meta", "LinkedIn", "TikTok", "Email"]
    weights = {"Google Ads": 0.34, "Meta": 0.26, "LinkedIn": 0.16, "TikTok": 0.14, "Email": 0.10}
    total_spend = 420_000
    channels = []
    for c in chans:
        spend = round(total_spend * weights[c] * rng.uniform(0.9, 1.1))
        roas = rng.uniform(2.1, 6.5) if c != "Email" else rng.uniform(5.0, 9.0)
        revenue = round(spend * roas)
        leads = round(spend / rng.uniform(18, 55))
        customers = max(1, round(leads * rng.uniform(0.08, 0.2)))
        cac = round(spend / customers)
        channels.append({"name": c, "spend": spend, "revenue": revenue,
                         "roas": round(roas, 2), "leads": leads, "cac": cac})
    tot_spend = sum(c["spend"] for c in channels)
    tot_rev = sum(c["revenue"] for c in channels)
    tot_leads = sum(c["leads"] for c in channels)
    # gentle, readable funnel taper (visitor-based) — reads cleanly on a dashboard
    visitors = round(tot_leads / 0.40)
    funnel = [
        {"name": "Visitors", "value": visitors},
        {"name": "Leads", "value": round(visitors * 0.40)},
        {"name": "MQLs", "value": round(visitors * 0.22)},
        {"name": "SQLs", "value": round(visitors * 0.11)},
        {"name": "Customers", "value": round(visitors * 0.045)},
    ]
    customers = funnel[-1]["value"]
    season = [0.85, 0.9, 0.95, 1.0, 1.05, 1.02, 0.98, 1.0, 1.08, 1.12, 1.2, 1.25]
    leads_month = [round(tot_leads / 12 * s) for s in season]
    kpis = [
        {"label": "Marketing Spend", "value": eurk(tot_spend), "accent": VIOLET, "delta": "YTD", "deltaUp": True, "sub": "all channels"},
        {"label": "Blended ROAS", "value": f"{tot_rev / tot_spend:.1f}x", "accent": GREEN, "delta": "+0.4x", "deltaUp": True, "sub": "on ad spend"},
        {"label": "Avg CAC", "value": eur(tot_spend / max(customers, 1)), "accent": AMBER, "delta": "-8%", "deltaUp": True, "sub": "per customer"},
        {"label": "Total Leads", "value": f"{tot_leads:,}", "accent": CYAN, "delta": "+14%", "deltaUp": True, "sub": "captured"},
        {"label": "Conversion Rate", "value": pct(customers / visitors * 100), "accent": BLUE, "delta": "+0.6 pts", "deltaUp": True, "sub": "visitor → customer"},
    ]
    data = {
        "id": "marketing", "slug": "marketing", "company": "Pulse Media",
        "title": "Marketing & Social Analytics Dashboard", "currency": "€", "period": "FY 2025",
        "months": MONTHS, "channels": [c["name"] for c in channels], "kpis": kpis,
        "series": {
            "channelPerformance": channels, "funnel": funnel, "leadsByMonth": leads_month,
            "roasByChannel": [{"name": c["name"], "value": c["roas"]} for c in channels],
            "spendByChannel": [{"name": c["name"], "value": c["spend"]} for c in channels],
        },
    }
    _excel_generic(data, "marketing", VIOLET, [
        ("bar", "ROAS by Channel", "roasByChannel", None, None),
        ("doughnut", "Spend by Channel", "spendByChannel", None, None),
        ("line", "Leads by Month", "leadsByMonth", None, MONTHS),
    ], ("Channel Performance", ["Channel", "Spend (€)", "Revenue (€)", "ROAS", "Leads", "CAC (€)"],
        [[c["name"], c["spend"], c["revenue"], c["roas"], c["leads"], c["cac"]] for c in channels], [1, 2, 5], None))
    return data


# ======================================================================================
# 7. HR — BrightPath
# ======================================================================================
def build_hr() -> dict[str, Any]:
    rng = random.Random(707)
    depts = ["Engineering", "Sales", "Marketing", "Operations", "Support", "People"]
    hc = {d: rng.randint(18, 120) for d in depts}
    total_hc = sum(hc.values())
    attr = {d: round(rng.uniform(6, 22), 1) for d in depts}
    avg_attr = sum(attr.values()) / len(depts)
    season = [1.4, 1.1, 1.0, 0.9, 1.2, 1.6, 1.3, 1.0, 0.8, 1.1, 1.0, 1.5]
    attr_month = [round(avg_attr / 12 * s * total_hc / 100, 1) for s in season]
    funnel = [
        {"name": "Applied", "value": rng.randint(1800, 2400)},
        {"name": "Screened", "value": rng.randint(700, 1000)},
        {"name": "Interviewed", "value": rng.randint(280, 420)},
        {"name": "Offer", "value": rng.randint(90, 140)},
        {"name": "Hired", "value": rng.randint(60, 95)},
    ]
    tenure = [
        {"name": "< 1 yr", "value": round(total_hc * 0.24)},
        {"name": "1–2 yrs", "value": round(total_hc * 0.31)},
        {"name": "2–4 yrs", "value": round(total_hc * 0.28)},
        {"name": "4+ yrs", "value": round(total_hc * 0.17)},
    ]
    kpis = [
        {"label": "Headcount", "value": f"{total_hc:,}", "accent": CYAN, "delta": f"+{funnel[-1]['value']} hires", "deltaUp": True, "sub": "full-time"},
        {"label": "Attrition Rate", "value": pct(avg_attr), "accent": RED, "delta": "-1.8 pts", "deltaUp": True, "sub": "annualised"},
        {"label": "Open Roles", "value": f"{rng.randint(14, 32)}", "accent": AMBER, "delta": "actively hiring", "deltaUp": True, "sub": "across teams"},
        {"label": "Avg Tenure", "value": "2.4 yrs", "accent": INDIGO, "delta": "+0.2", "deltaUp": True, "sub": "company-wide"},
        {"label": "Employee eNPS", "value": "+38", "accent": GREEN, "delta": "+5", "deltaUp": True, "sub": "engagement"},
    ]
    data = {
        "id": "hr", "slug": "hr", "company": "BrightPath",
        "title": "HR & People Analytics Dashboard", "currency": "€", "period": "FY 2025",
        "months": MONTHS, "departments": depts, "kpis": kpis,
        "series": {
            "headcountByDept": [{"name": d, "value": hc[d]} for d in depts],
            "attritionByMonth": attr_month, "hiringFunnel": funnel, "tenureDistribution": tenure,
        },
        "table": [{"name": d, "headcount": hc[d], "attrition": attr[d]} for d in depts],
    }
    _excel_generic(data, "hr", CYAN, [
        ("bar", "Headcount by Department", "headcountByDept", None, None),
        ("doughnut", "Tenure Distribution", "tenureDistribution", None, None),
        ("line", "Attrition (leavers) by Month", "attritionByMonth", None, MONTHS),
    ], ("Department People KPIs", ["Department", "Headcount", "Attrition %"],
        [[r["name"], r["headcount"], r["attrition"]] for r in data["table"]], [], None))
    return data


# ======================================================================================
# 8. E-COMMERCE — Kavo Store
# ======================================================================================
def build_ecommerce() -> dict[str, Any]:
    rng = random.Random(808)
    season = [0.82, 0.8, 0.9, 0.95, 1.0, 1.02, 0.98, 0.96, 1.06, 1.14, 1.32, 1.4]
    revenue = [round(190_000 * s * rng.uniform(0.95, 1.05)) for s in season]
    orders = [round(revenue[i] / rng.uniform(58, 82)) for i in range(12)]
    total_rev = sum(revenue)
    total_orders = sum(orders)
    aov = total_rev / total_orders
    sources = ["Organic", "Paid Search", "Social", "Email", "Direct"]
    sw = {"Organic": 0.32, "Paid Search": 0.26, "Social": 0.18, "Email": 0.14, "Direct": 0.10}
    src_perf = []
    for s in sources:
        rev = round(total_rev * sw[s] * rng.uniform(0.9, 1.1))
        sessions = round(rev / rng.uniform(4, 9))
        conv = rng.uniform(1.4, 4.2)
        src_perf.append({"name": s, "revenue": rev, "sessions": sessions, "conversion": round(conv, 2)})
    products = ["Nomad Backpack", "Aero Runners", "Lume Bottle", "Terra Tee",
               "Pulse Buds", "Cove Hoodie", "Drift Cap", "Halo Ring Light"]
    top_products = sorted(
        [{"name": p, "value": round(total_rev * rng.uniform(0.03, 0.09))} for p in products],
        key=lambda x: -x["value"])[:6]
    conv_rate = sum(s["conversion"] for s in src_perf) / len(src_perf)
    kpis = [
        {"label": "Revenue", "value": eurk(total_rev), "accent": VIOLET, "delta": "+18%", "deltaUp": True, "sub": "YoY"},
        {"label": "Orders", "value": f"{total_orders:,}", "accent": INDIGO, "delta": "+12%", "deltaUp": True, "sub": "full year"},
        {"label": "Avg Order Value", "value": eur(aov), "accent": CYAN, "delta": "+4.5%", "deltaUp": True, "sub": "per order"},
        {"label": "Conversion Rate", "value": pct(conv_rate), "accent": GREEN, "delta": "+0.3 pts", "deltaUp": True, "sub": "blended"},
        {"label": "Repeat Customers", "value": "38.2%", "accent": AMBER, "delta": "+2.1 pts", "deltaUp": True, "sub": "of orders"},
    ]
    data = {
        "id": "ecommerce", "slug": "ecommerce", "company": "Kavo Store",
        "title": "E-commerce Analytics Dashboard", "currency": "€", "period": "FY 2025",
        "months": MONTHS, "sources": sources, "kpis": kpis,
        "series": {
            "revenueByMonth": revenue, "ordersByMonth": orders,
            "revenueBySource": [{"name": s["name"], "value": s["revenue"]} for s in src_perf],
            "topProducts": top_products, "sourcePerformance": src_perf,
        },
    }
    _excel_generic(data, "ecommerce", INDIGO, [
        ("combo", "Revenue vs Orders by Month", "revenueByMonth", "ordersByMonth", MONTHS),
        ("doughnut", "Revenue by Source", "revenueBySource", None, None),
        ("bar", "Top Products", "topProducts", None, None),
    ], ("Traffic Source Performance", ["Source", "Revenue (€)", "Sessions", "Conversion %"],
        [[s["name"], s["revenue"], s["sessions"], s["conversion"]] for s in src_perf], [1, 2], None))
    return data


# --------------------------------------------------------------------------------------
# Generic Excel builder (used by the 4 newer dashboards)
# --------------------------------------------------------------------------------------
def _excel_generic(data, name, accent, charts, table):
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    set_widths(dash, {i: 11 for i in range(1, 17)})
    title_band(dash, data["company"], data["title"])
    for i, k in enumerate(data["kpis"][:5]):
        kpi_card(dash, 6, 1 + i * 3, k["label"], k["value"], k["accent"])

    md = wb.create_sheet("Model")
    md.sheet_state = "hidden"
    col = 1
    ref = {}
    for spec in charts:
        kind, title_txt, key1, key2, cats = spec
        series = data["series"][key1]
        if isinstance(series, list) and series and isinstance(series[0], dict):
            names = [x["name"] for x in series]
            vals = [x["value"] for x in series]
        else:
            names = cats
            vals = series
        md.cell(row=1, column=col, value="cat")
        md.cell(row=1, column=col + 1, value=title_txt)
        for r, (nm, vv) in enumerate(zip(names, vals)):
            md.cell(row=2 + r, column=col, value=nm)
            md.cell(row=2 + r, column=col + 1, value=vv)
        ref[key1] = (col, len(vals))
        if key2:
            s2 = data["series"][key2]
            md.cell(row=1, column=col + 2, value=key2)
            for r, vv in enumerate(s2):
                md.cell(row=2 + r, column=col + 2, value=vv)
            ref[key2] = (col + 2, len(s2))
        col += 4

    anchors = ["A10", "H10", "N10"]
    for idx, spec in enumerate(charts):
        kind, title_txt, key1, key2, cats = spec
        c0, n = ref[key1]
        if kind == "doughnut":
            ch = DoughnutChart()
        elif kind == "line":
            ch = LineChart()
        else:
            ch = BarChart()
            ch.type = "col"
        ch.title = title_txt
        ch.height = 7.5
        ch.width = 8.5
        ch.add_data(Reference(md, min_col=c0 + 1, min_row=1, max_row=1 + n), titles_from_data=True)
        ch.set_categories(Reference(md, min_col=c0, min_row=2, max_row=1 + n))
        if kind == "doughnut":
            color_pie_points(ch, n)
        elif kind == "combo" and key2:
            ch.series[0].graphicalProperties.solidFill = accent
            c2, n2 = ref[key2]
            line = LineChart()
            line.add_data(Reference(md, min_col=c2 + 1, min_row=1, max_row=1 + n2), titles_from_data=True)
            line.series[0].graphicalProperties.line.solidFill = CYAN
            line.series[0].graphicalProperties.line.width = 26000
            ch += line
        else:
            ch.series[0].graphicalProperties.solidFill = accent
            if kind == "line":
                ch.series[0].graphicalProperties.line.solidFill = accent
                ch.series[0].graphicalProperties.line.width = 26000
            ch.legend = None
        dash.add_chart(ch, anchors[idx])

    if table:
        sheet_name, headers, rows, num_cols, cond_col = table
        ds = wb.create_sheet(sheet_name[:31])
        ds.sheet_view.showGridLines = False
        for c, h in enumerate(headers, start=1):
            ds.cell(row=1, column=c, value=h)
        for i, row in enumerate(rows, start=2):
            for c, v in enumerate(row, start=1):
                ds.cell(row=i, column=c, value=v)
            for nc in num_cols:
                ds.cell(row=i, column=nc + 1).number_format = "#,##0"
        last = len(rows) + 1
        style_table_sheet(ds, 1, 2, last, len(headers), sheet_name.replace(" ", ""))
        set_widths(ds, {1: 20, **{c: 13 for c in range(2, len(headers) + 1)}})
        if cond_col is not None:
            ds.conditional_formatting.add(
                f"{get_column_letter(cond_col + 1)}2:{get_column_letter(cond_col + 1)}{last}",
                ColorScaleRule(start_type="min", start_color=RED,
                               mid_type="percentile", mid_value=50, mid_color=AMBER,
                               end_type="max", end_color=GREEN))
    _finish(wb, name)


# --------------------------------------------------------------------------------------
# Finalise
# --------------------------------------------------------------------------------------
def _finish(wb: Workbook, name: str) -> None:
    path = os.path.join(XLSX_DIR, f"{name}-dashboard.xlsx")
    wb.save(path)
    print(f"  xlsx  -> {os.path.relpath(path, ROOT)}")


def main() -> None:
    print("Generating Vexloft Data demo dashboards...")
    builders = {
        "sales": build_sales,
        "inventory": build_inventory,
        "procurement": build_procurement,
        "executive": build_executive,
        "financial": build_financial,
        "marketing": build_marketing,
        "hr": build_hr,
        "ecommerce": build_ecommerce,
    }
    index = []
    for name, fn in builders.items():
        data = fn()
        json_path = os.path.join(JSON_DIR, f"{name}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  json  -> {os.path.relpath(json_path, ROOT)}")
        index.append({
            "slug": data["slug"], "company": data["company"],
            "title": data["title"], "period": data["period"],
        })
    with open(os.path.join(JSON_DIR, "index.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    print("Done.")


if __name__ == "__main__":
    main()
